from openpyxl import load_workbook
from unidecode import unidecode
from datetime import datetime, time


class XlsxEditor():
    
    def __init__(self, xlsxFilePath):
        self.xlsxFilePath = xlsxFilePath
        # Abre o arquivo e recebe dicionário com todas as tabelas
        self.wb = load_workbook(filename=xlsxFilePath)

    def rowFinder(self, nome):
        now = datetime.now()
        self.wb.active = self.wb.get_sheet_by_name(nome)
        ws = self.wb.active
        # Percorre linhas para encontrar dia de hoje
        # Procura na coluna "B", onde estão as datas
        for i in range(10, (ws.max_row)+1):
            if ws.cell(row=i, column=2).value == datetime(now.year, now.month, now.day):
                return i, ws

    def insertData(self, nome, evento):
        linha, ws = self.rowFinder(nome)
        coluna = {'entrada': 3, 'inicio_intervalo': 4, 'fim_intervalo': 5, 'saida': 6}

        ws.cell(row=linha, column=coluna[evento]).value = datetime.now().strftime("%H:%M:00")
        self.writeExcel()

        # Retorna dados preenchidos hoje
        return [str(c.value)[:5] for c in ws[linha][2:6]]

    def writeExcel(self):
        self.wb.save(self.xlsxFilePath)


if __name__ == "__main__":
    try:
        dados = XlsxEditor('Eduardo', 'saida')
        print(dados.atual)
    except RuntimeError as e:
        print(e)

            